#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename="student.xlsx")
ws = wb.active

row_max = ws.max_row

for  row  in  range(2, row_max+1):
    ws['G' + str(row)].value = ws['C' + str(row)].value * 0.3 + ws['D' + str(row)].value * 0.35 + ws['E' + str(row)].value * 0.34 + ws['F' + str(row)].value

lst = []
lst1 = []
lst2 = []

for cell in ws['A']:   # 학번 저장
    lst1.append(cell.value)

for cell in ws['G']:   # total 저장
    lst2.append(cell.value)

del lst1[0]   # 'id' 삭제
del lst2[0]   # 'total' 삭제

dic = dict(zip(lst2,lst1))

lst2.sort(reverse=True)

a_len = int(row_max * 0.3)
b_len = int(row_max * 0.7)
A = lst2[:a_len]
B = lst2[a_len:b_len]
other = lst2[b_len:]

for i in range(a_len):
    if i < int(a_len * 0.5):
        li = [dic[A[i]], A[i], 'A+']
    else:
        li = [dic[A[i]], A[i], 'A0']
    lst.append(li)

for i in range(len(B)):
    if i < int(len(B) * 0.5):
        li = [dic[B[i]], B[i], 'B+']
    else:
        li = [dic[B[i]], B[i], 'B0']
    lst.append(li)

for i in range(len(other)):
    if other[i] < 40:
        li = [dic[other[i]], other[i], 'F']
    elif i < int(len(other) * 0.5):
        li = [dic[other[i]], other[i], 'C+']
    else:
        li = [other[B[i]], other[i], 'C0']
    lst.append(li)

lst.sort()

for  row  in  range(2, row_max+1):
    ws['H' + str(row)].value = lst[row-2][2]

wb.save('student.xlsx')


#커밋 테스트